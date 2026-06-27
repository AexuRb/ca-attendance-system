import argparse
import os
import re
import sys
from pathlib import Path

import bcrypt
import mysql.connector
from openpyxl import load_workbook


def clean(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_grade(value):
    text = clean(value)
    match = re.search(r"(\d{4})\s*级", text)
    if match:
        return f"{match.group(1)}级"
    return text


def find_header_row(ws):
    for row_no, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        headers = [clean(cell) for cell in row]
        if any("学号" in header for header in headers) and any("姓名" in header for header in headers):
            return row_no, read_columns(headers)
    return None, fallback_columns()


def read_columns(headers):
    columns = {}
    for index, header in enumerate(headers):
        text = header.lower()
        if "学号" in text:
            columns.setdefault("student_no", index)
        if "姓名" in text:
            columns.setdefault("name", index)
        if "年级" in text:
            columns.setdefault("grade", index)

    if "student_no" not in columns or "name" not in columns:
        return fallback_columns()

    columns.setdefault("grade", -1)
    return columns


def fallback_columns():
    return {
        "name": 1,
        "grade": 4,
        "student_no": 5,
    }


def row_value(row, columns, key):
    index = columns.get(key, -1)
    if index is None or index < 0 or index >= len(row):
        return ""
    return row[index]


def parse_members(path):
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb.active
        members = []
        skipped = []
        seen = set()
        duplicates = []
        header_row, columns = find_header_row(ws)
        start_row = header_row + 1 if header_row else 3

        for row_no, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
            name = clean(row_value(row, columns, "name"))
            grade = normalize_grade(row_value(row, columns, "grade"))
            student_no = clean(row_value(row, columns, "student_no"))

            if not any([name, student_no, grade]):
                continue
            if not name or not student_no:
                skipped.append((row_no, "missing_name_or_student_no"))
                continue
            if not re.fullmatch(r"\d{6,32}", student_no):
                skipped.append((row_no, f"invalid_student_no:{student_no}"))
                continue
            if student_no in seen:
                duplicates.append((row_no, student_no))
                continue
            seen.add(student_no)

            members.append(
                {
                    "student_no": student_no,
                    "name": name,
                    "grade": grade or None,
                }
            )

        return members, skipped, duplicates
    finally:
        wb.close()


def make_password_hash(student_no):
    initial_password = student_no[-6:]
    return bcrypt.hashpw(initial_password.encode("utf-8"), bcrypt.gensalt(rounds=10)).decode("utf-8")


def connect(args):
    password = args.password or os.environ.get("MYSQL_PWD") or os.environ.get("DB_PASSWORD")
    if not password:
        raise RuntimeError("Missing MySQL password. Pass --password or set MYSQL_PWD/DB_PASSWORD.")
    return mysql.connector.connect(
        host=args.host,
        port=args.port,
        user=args.user,
        password=password,
        database=args.database,
        charset="utf8mb4",
        use_unicode=True,
    )


def import_members(args, members):
    sql = """
        INSERT INTO users (
          student_no, name, password_hash, role, status, grade, must_change_password
        ) VALUES (
          %(student_no)s, %(name)s, %(password_hash)s, 'MEMBER', 'ACTIVE',
          %(grade)s, 1
        )
        ON DUPLICATE KEY UPDATE
          name = VALUES(name),
          grade = VALUES(grade),
          updated_at = CURRENT_TIMESTAMP
    """
    payload = []
    for member in members:
        item = dict(member)
        item["password_hash"] = make_password_hash(member["student_no"])
        payload.append(item)

    conn = connect(args)
    try:
        cur = conn.cursor()
        cur.executemany(sql, payload)
        conn.commit()
        cur.close()
    finally:
        conn.close()


def main():
    parser = argparse.ArgumentParser(description="Import association members from Excel into MySQL.")
    parser.add_argument("--excel", default=os.environ.get("MEMBER_EXCEL"), help="Path to the member Excel file.")
    parser.add_argument("--host", default="127.0.0.1")
    parser.add_argument("--port", type=int, default=3306)
    parser.add_argument("--user", default="root")
    parser.add_argument("--password", default=None)
    parser.add_argument("--database", default="ca_attendance")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    if not args.excel:
        print("Missing Excel file. Pass --excel or set MEMBER_EXCEL.", file=sys.stderr)
        return 1

    path = Path(args.excel)
    if not path.exists():
        print(f"Excel file not found: {path}", file=sys.stderr)
        return 1

    members, skipped, duplicates = parse_members(path)
    print(f"valid_members={len(members)}")
    print(f"skipped_rows={len(skipped)}")
    print(f"duplicate_rows={len(duplicates)}")
    if skipped[:10]:
        print("skipped_sample=" + repr(skipped[:10]))
    if duplicates[:10]:
        print("duplicate_sample=" + repr(duplicates[:10]))

    if args.dry_run:
        return 0

    import_members(args, members)
    print(f"imported_or_updated={len(members)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
